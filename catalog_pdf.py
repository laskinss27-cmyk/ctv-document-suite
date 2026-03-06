"""
CTV PDF Generator — ядро генерации PDF из xlsx-файла коммерческого предложения.
"""

import zipfile, re, os, io
from pathlib import Path

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (Paragraph, Spacer, Table, TableStyle,
                                 HRFlowable, PageBreak, BaseDocTemplate,
                                 PageTemplate, Frame, NextPageTemplate,
                                 Image as RLImage)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from PIL import Image as PILImage
import openpyxl

W, H = A4

# ── Colors ───────────────────────────────────────────────────────────────────
DARK_BLUE   = colors.HexColor('#1A2B4A')
ACCENT_BLUE = colors.HexColor('#2563EB')
LIGHT_BLUE  = colors.HexColor('#EFF6FF')
GOLD        = colors.HexColor('#F59E0B')
GRAY        = colors.HexColor('#64748B')
WHITE       = colors.white
DARK_TEXT   = colors.HexColor('#1E293B')

def _register_fonts():
    """Регистрирует шрифты с поддержкой кириллицы."""
    candidates = [
        '/usr/share/fonts/truetype/dejavu/',
        '/usr/share/fonts/dejavu/',
        '/Library/Fonts/',
        'C:/Windows/Fonts/',
    ]
    for path in candidates:
        reg = Path(path) / 'DejaVuSans.ttf'
        bold = Path(path) / 'DejaVuSans-Bold.ttf'
        ital = Path(path) / 'DejaVuSans-Oblique.ttf'
        if reg.exists():
            pdfmetrics.registerFont(TTFont('DV',  str(reg)))
            pdfmetrics.registerFont(TTFont('DVB', str(bold)))
            pdfmetrics.registerFont(TTFont('DVI', str(ital)))
            return True
    # Fallback — попробуем установить
    try:
        import subprocess
        subprocess.run(['apt-get', 'install', '-y', 'fonts-dejavu'],
                       capture_output=True)
        p = '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'
        if Path(p).exists():
            pdfmetrics.registerFont(TTFont('DV',  p))
            pdfmetrics.registerFont(TTFont('DVB', p.replace('Sans.', 'Sans-Bold.')))
            pdfmetrics.registerFont(TTFont('DVI', p.replace('Sans.', 'Sans-Oblique.')))
            return True
    except Exception:
        pass
    return False


def _extract_xlsx_data(xlsx_path: str):
    """
    Извлекает товары и изображения из xlsx-файла.
    Возвращает (header_info, products, logo_img_bytes).
    """
    xlsx_path = str(xlsx_path)

    # ── Images from zip ──────────────────────────────────────────────────────
    rid_to_bytes = {}
    logo_bytes = None
    row_to_img_bytes = {}

    with zipfile.ZipFile(xlsx_path, 'r') as z:
        files = z.namelist()
        media = {os.path.basename(f): z.read(f) for f in files if 'media' in f}

        # relationships
        rels_raw = z.read('xl/drawings/_rels/drawing1.xml.rels').decode()
        draw_raw = z.read('xl/drawings/drawing1.xml').decode()

    rid_to_fname = dict(re.findall(
        r'Id="(rId\d+)"[^>]+Target="\.\./media/([^"]+)"', rels_raw))

    anchors = re.findall(
        r'<xdr:from>\s*<xdr:col>(\d+)</xdr:col>.*?'
        r'<xdr:row>(\d+)</xdr:row>.*?r:embed="(rId\d+)"',
        draw_raw, re.DOTALL)

    for col, row, rid in anchors:
        fname = rid_to_fname.get(rid, '')
        img_bytes = media.get(fname)
        if not img_bytes:
            continue
        if col == '0':          # логотип компании
            logo_bytes = img_bytes
        elif col == '1':        # фото товара
            row_to_img_bytes[int(row)] = img_bytes

    # ── Xlsx data ────────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    header = {
        'company': str(ws['A2'].value or ''),
        'manager': str(ws['C2'].value or ''),
        'title':   str(ws['A3'].value or 'Коммерческое предложение'),
    }

    products = []
    data_start = 7          # строка Excel (1-indexed), где начинаются товары
    for i, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True)):
        if row[0] is None or not isinstance(row[0], int):
            break
        art, _, name, desc, price, qty, total = row
        draw_row = (data_start - 1) + i   # 0-indexed drawing row
        products.append({
            'article':   art,
            'name':      name or '',
            'desc':      desc or '',
            'price':     price or 0,
            'qty':       qty or 1,
            'total':     total or 0,
            'img_bytes': row_to_img_bytes.get(draw_row),
        })

    return header, products, logo_bytes


def _rl_image_from_bytes(img_bytes, max_w, max_h):
    """Создаёт ReportLab Image из байт, масштабируя под max_w × max_h."""
    if not img_bytes:
        return None
    try:
        pil = PILImage.open(io.BytesIO(img_bytes))
        pw, ph = pil.size
        scale = min(max_w / pw, max_h / ph)
        nw, nh = pw * scale, ph * scale
        buf = io.BytesIO(img_bytes)
        return RLImage(buf, width=nw, height=nh)
    except Exception:
        return None


def _style(name, **kw):
    d = dict(fontName='DV', fontSize=9, textColor=DARK_TEXT, leading=13)
    d.update(kw)
    return ParagraphStyle(name, **d)


def _make_cover(header, logo_bytes, products):
    """Возвращает функцию-callback для обложки."""
    def _draw(canvas, doc):
        canvas.saveState()

        # Фон
        canvas.setFillColor(DARK_BLUE)
        canvas.rect(0, 0, W, H, fill=1, stroke=0)
        canvas.setFillColor(colors.HexColor('#1E3A6E'))
        canvas.circle(W - 50*mm, H - 20*mm, 130*mm, fill=1, stroke=0)
        canvas.setFillColor(colors.HexColor('#152540'))
        canvas.circle(-20*mm, 70*mm, 100*mm, fill=1, stroke=0)

        # Полосы
        canvas.setFillColor(GOLD)
        canvas.rect(0, H - 8*mm, W, 8*mm, fill=1, stroke=0)
        canvas.setFillColor(ACCENT_BLUE)
        canvas.rect(0, 0, W, 6*mm, fill=1, stroke=0)

        # Логотип
        if logo_bytes:
            try:
                pil = PILImage.open(io.BytesIO(logo_bytes))
                pw, ph = pil.size
                scale = min(40*mm / pw, 18*mm / ph)
                buf = io.BytesIO(logo_bytes)
                canvas.drawImage(buf, 20*mm, H - 35*mm,
                                 width=pw*scale, height=ph*scale,
                                 preserveAspectRatio=True, mask='auto')
            except Exception:
                pass

        canvas.setFillColor(colors.HexColor('#94A3B8'))
        canvas.setFont('DV', 8)
        canvas.drawString(20*mm, H - 38*mm, 'Системы безопасности и домофонии')
        canvas.setStrokeColor(colors.HexColor('#2D4A7A'))
        canvas.line(20*mm, H - 43*mm, W - 20*mm, H - 43*mm)

        # Заголовок
        canvas.setFillColor(WHITE)
        canvas.setFont('DVB', 34)
        canvas.drawString(20*mm, H/2 + 42*mm, 'КАТАЛОГ')
        canvas.drawString(20*mm, H/2 + 24*mm, 'ОБОРУДОВАНИЯ')
        canvas.setFillColor(GOLD)
        canvas.setFont('DVB', 15)
        canvas.drawString(20*mm, H/2 + 10*mm, header['title'])
        canvas.setFillColor(ACCENT_BLUE)
        canvas.rect(20*mm, H/2 + 4*mm, 50*mm, 2*mm, fill=1, stroke=0)

        # Блок с инфо о документе
        canvas.setFillColor(colors.HexColor('#1E3A6E'))
        canvas.roundRect(20*mm, H/2 - 30*mm, W - 40*mm, 28*mm, 6, fill=1, stroke=0)
        canvas.setFillColor(WHITE)
        canvas.setFont('DVB', 10)
        canvas.drawString(26*mm, H/2 - 8*mm, header['title'])
        canvas.setFont('DV', 8.5)
        canvas.setFillColor(colors.HexColor('#94A3B8'))
        for i, line in enumerate(header['company'].split('\n')):
            canvas.drawString(26*mm, H/2 - 18*mm - i*11, line.strip())
        for i, line in enumerate(header['manager'].split('\n')):
            canvas.drawString(W/2, H/2 - 18*mm - i*11, line.strip())

        # Статистика
        total_sum = sum(p['total'] for p in products)
        stats = [(str(len(products)), 'позиций'),
                 (f'{total_sum:,}'.replace(',', ' '), 'руб. итого')]
        for i, (val, lbl) in enumerate(stats):
            x = 20*mm + i * 85*mm
            canvas.setFillColor(colors.HexColor('#0F1F3D'))
            canvas.roundRect(x, H/2 - 65*mm, 78*mm, 28*mm, 5, fill=1, stroke=0)
            canvas.setFillColor(GOLD)
            canvas.setFont('DVB', 15)
            canvas.drawCentredString(x + 39*mm, H/2 - 44*mm, val)
            canvas.setFillColor(colors.HexColor('#94A3B8'))
            canvas.setFont('DV', 8)
            canvas.drawCentredString(x + 39*mm, H/2 - 55*mm, lbl)

        canvas.setFillColor(colors.HexColor('#64748B'))
        canvas.setFont('DV', 8)
        canvas.drawString(20*mm, 15*mm,
                          'Цены действительны в течение 3 банковских дней')
        canvas.restoreState()

    return _draw


def _make_page_bg(canvas, doc):
    canvas.saveState()
    canvas.setFillColor(WHITE)
    canvas.rect(0, 0, W, H, fill=1, stroke=0)
    canvas.setFillColor(DARK_BLUE)
    canvas.rect(0, H - 15*mm, W, 15*mm, fill=1, stroke=0)
    canvas.setFillColor(GOLD)
    canvas.rect(0, H - 16.5*mm, W, 1.5*mm, fill=1, stroke=0)
    canvas.setFillColor(WHITE)
    canvas.setFont('DVB', 10)
    canvas.drawString(20*mm, H - 10*mm, 'Коммерческое предложение • CTV')
    canvas.setFont('DV', 9)
    canvas.drawRightString(W - 20*mm, H - 10*mm, f'Стр. {doc.page}')
    canvas.setFillColor(DARK_BLUE)
    canvas.rect(0, 0, W, 10*mm, fill=1, stroke=0)
    canvas.setFillColor(colors.HexColor('#64748B'))
    canvas.setFont('DV', 7)
    canvas.drawCentredString(W/2, 3.5*mm,
                             'CTV • Системы безопасности и домофонии')
    canvas.restoreState()


def _build_product_card(product):
    is_monitor = 'монитор' in product['name'].lower()
    accent  = ACCENT_BLUE if is_monitor else GOLD
    card_bg = LIGHT_BLUE  if is_monitor else colors.HexColor('#FFFBEB')
    cat_txt = 'МОНИТОР ВИДЕОДОМОФОНА' if is_monitor else 'ВЫЗЫВНАЯ ВИДЕОПАНЕЛЬ'

    desc_lines = [l.strip() for l in product['desc'].split('\n') if l.strip()]
    spec_lines = [l for l in desc_lines if '\t' in l]
    narr_lines = [l for l in desc_lines if '\t' not in l]
    narrative  = ' '.join(narr_lines[:3])

    img = _rl_image_from_bytes(product['img_bytes'], 38*mm, 50*mm)
    img_cell = img if img else Spacer(38*mm, 50*mm)

    cat_p  = Paragraph(cat_txt, _style('cat', fontName='DVB', fontSize=7,
                                        textColor=WHITE, backColor=accent,
                                        borderPadding=(2,5,2,5)))
    name_p = Paragraph(product['name'], _style('nm', fontName='DVB', fontSize=11,
                                                textColor=DARK_BLUE, leading=14))
    art_p  = Paragraph(f'Арт. {product["article"]}',
                        _style('art', fontSize=8, textColor=GRAY))
    narr_p = (Paragraph(narrative, _style('narr', fontName='DVI', fontSize=8.5,
                                           textColor=GRAY, leading=12))
              if narrative else Spacer(1, 2*mm))

    right = Table([[cat_p], [Spacer(1,1*mm)], [name_p], [art_p],
                   [Spacer(1,1*mm)], [narr_p]],
                  colWidths=[115*mm])
    right.setStyle(TableStyle([
        ('VALIGN',        (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('RIGHTPADDING',  (0,0), (-1,-1), 0),
        ('TOPPADDING',    (0,0), (-1,-1), 1),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1),
    ]))

    main = Table([[img_cell, right]], colWidths=[42*mm, 120*mm])
    main.setStyle(TableStyle([
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND',    (0,0), (-1,-1), card_bg),
        ('TOPPADDING',    (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING',   (0,0), (0,-1), 8),
        ('LEFTPADDING',   (1,0), (1,-1), 6),
        ('RIGHTPADDING',  (0,0), (-1,-1), 8),
        ('ROUNDEDCORNERS', [6,6,6,6]),
    ]))

    spec_rows = []
    for line in spec_lines[:7]:
        parts = line.split('\t', 1)
        if len(parts) == 2:
            k, v = parts
            spec_rows.append([
                Paragraph(k.strip(), _style('sk', fontSize=7.5,
                                             textColor=GRAY, leading=11)),
                Paragraph(v.strip(), _style('sv', fontSize=7.5, fontName='DVB',
                                             textColor=DARK_TEXT, leading=11)),
            ])

    price_txt = f'{product["price"]:,}'.replace(',', ' ')
    bar = Table([[Paragraph(f'Цена: <b>{price_txt} руб.</b>',
                             _style('pb', fontSize=11, textColor=WHITE,
                                    alignment=TA_CENTER, leading=14))]],
                colWidths=[165*mm])
    bar.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), accent),
        ('TOPPADDING',    (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('ROUNDEDCORNERS', [4,4,4,4]),
    ]))

    els = [main, Spacer(1, 2*mm)]
    if spec_rows:
        st = Table(spec_rows, colWidths=[70*mm, 90*mm])
        st.setStyle(TableStyle([
            ('VALIGN',        (0,0), (-1,-1), 'TOP'),
            ('TOPPADDING',    (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ('LEFTPADDING',   (0,0), (-1,-1), 4),
            ('ROWBACKGROUNDS',(0,0), (-1,-1),
             [colors.HexColor('#F8FAFC'), WHITE]),
        ]))
        els += [st, Spacer(1, 2*mm)]
    els += [bar, Spacer(1, 5*mm),
            HRFlowable(width='100%', thickness=0.5,
                       color=colors.HexColor('#E2E8F0'), spaceAfter=5*mm)]
    return els


def _section_header(title, subtitle, line_color):
    h = ParagraphStyle('sh', fontName='DVB', fontSize=16, textColor=DARK_BLUE,
                        leading=20, spaceAfter=2*mm)
    s = ParagraphStyle('ss', fontName='DVI', fontSize=9, textColor=GRAY,
                        leading=13, spaceAfter=4*mm)
    return [Paragraph(title, h), Paragraph(subtitle, s),
            HRFlowable(width='100%', thickness=2,
                       color=line_color, spaceAfter=5*mm)]


class _MyDoc(BaseDocTemplate):
    def __init__(self, filename, cover_fn, **kw):
        BaseDocTemplate.__init__(self, filename, **kw)
        self.addPageTemplates([
            PageTemplate(id='Cover',
                         frames=[Frame(0, 0, W, H,
                                       leftPadding=0, rightPadding=0,
                                       topPadding=0, bottomPadding=0)],
                         onPage=cover_fn),
            PageTemplate(id='Content',
                         frames=[Frame(20*mm, 15*mm, W-40*mm, H-33*mm,
                                       leftPadding=0, rightPadding=0,
                                       topPadding=0, bottomPadding=0)],
                         onPage=_make_page_bg),
        ])


def generate(xlsx_path_or_data, output_path: str, progress_cb=None) -> str:
    """
    xlsx_path_or_data: путь к xlsx (str) или готовый dict из parser.parse().
    """
    def _progress(step, total, msg):
        if progress_cb:
            progress_cb(step, total, msg)

    _progress(0, 5, 'Регистрация шрифтов…')
    _register_fonts()

    if isinstance(xlsx_path_or_data, dict):
        d = xlsx_path_or_data
        logo_bytes = d.get('logo_bytes')
        header = {'title': d.get('title',''), 'company': d.get('company',''),
                  'manager': d.get('manager','')}
        products = [{'article': p.get('article',''), 'name': p.get('name',''),
                     'desc': p.get('desc',''),
                     'price': float(p.get('price',0)), 'qty': float(p.get('qty',1)),
                     'total': float(p.get('total',0)) or float(p.get('price',0))*float(p.get('qty',1)),
                     'img_bytes': p.get('img_bytes')} for p in d.get('equipment',[])]
    else:
        _progress(1, 5, 'Чтение данных из файла…')
        header, products, logo_bytes = _extract_xlsx_data(xlsx_path_or_data)

    _progress(2, 5, f'Найдено {len(products)} позиций. Формирование страниц…')

    monitors = [p for p in products if 'монитор' in p['name'].lower()]
    panels   = [p for p in products if 'монитор' not in p['name'].lower()]

    story = [Spacer(1, H), NextPageTemplate('Content'), PageBreak()]

    if monitors:
        story += _section_header(
            'Мониторы видеодомофонов',
            'Цветные IPS-мониторы с поддержкой AHD/CVBS, Hands Free и записью на SD-карту',
            ACCENT_BLUE)
        for p in monitors:
            story += _build_product_card(p)

    if panels:
        story += _section_header(
            'Вызывные видеопанели',
            'Антивандальные панели с Full HD камерой, ИК-подсветкой и управлением замком',
            GOLD)
        for p in panels:
            story += _build_product_card(p)

    # Итоговая таблица
    story.append(PageBreak())
    story += _section_header('Итоговая таблица',
                              'Сводный прайс по всем позициям', ACCENT_BLUE)

    th = _style('th', fontName='DVB', fontSize=8.5, textColor=WHITE, leading=12)
    td = _style('td', fontSize=8, leading=12)
    tp = _style('tp', fontName='DVB', fontSize=8.5,
                textColor=ACCENT_BLUE, leading=12, alignment=TA_RIGHT)

    tdata = [[Paragraph('№', th), Paragraph('Наименование', th),
              Paragraph('Арт.', th), Paragraph('Цена', th)]]
    for i, p in enumerate(products, 1):
        price_txt = f'{p["price"]:,}'.replace(',', ' ') + ' руб.'
        tdata.append([Paragraph(str(i), td),
                      Paragraph(p['name'], td),
                      Paragraph(str(p['article']), td),
                      Paragraph(price_txt, tp)])

    row_colors = [('BACKGROUND', (0, i), (-1, i),
                   LIGHT_BLUE if i % 2 == 1 else WHITE)
                  for i in range(1, len(tdata))]
    sum_tbl = Table(tdata, colWidths=[8*mm, 103*mm, 20*mm, 34*mm])
    sum_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), DARK_BLUE),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING',   (0,0), (-1,-1), 5),
        ('RIGHTPADDING',  (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#E2E8F0')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ] + row_colors))
    story.append(sum_tbl)
    story.append(Spacer(1, 5*mm))

    total_sum = sum(p['total'] for p in products)
    tot_txt = f'{total_sum:,}'.replace(',', ' ')
    tot_tbl = Table(
        [[Paragraph(f'ИТОГО: {tot_txt} руб.',
                    _style('tot', fontName='DVB', fontSize=13,
                           textColor=WHITE, alignment=TA_CENTER, leading=18))]],
        colWidths=[165*mm])
    tot_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), DARK_BLUE),
        ('TOPPADDING',    (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('ROUNDEDCORNERS', [6,6,6,6]),
    ]))
    story.append(tot_tbl)
    story.append(Spacer(1, 6*mm))

    note_tbl = Table(
        [[Paragraph(
            'Цены действительны в течение 3 банковских дней.\n'
            f'По вопросам заказа: {header["manager"].replace(chr(10), " | ")}',
            _style('note', fontSize=8.5, textColor=WHITE,
                   alignment=TA_CENTER, leading=13))]],
        colWidths=[165*mm])
    note_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), colors.HexColor('#1E3A6E')),
        ('TOPPADDING',    (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('ROUNDEDCORNERS', [6,6,6,6]),
    ]))
    story.append(note_tbl)

    _progress(3, 5, 'Генерация PDF…')
    cover_fn = _make_cover(header, logo_bytes, products)
    doc = _MyDoc(output_path, cover_fn=cover_fn, pagesize=A4)
    doc.build(story)

    _progress(5, 5, f'Готово! Сохранён: {output_path}')
    return output_path
